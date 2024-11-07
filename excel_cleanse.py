#!/usr/bin/env python
# -*- coding: utf-8 -*-
# File_Name: excel_cleanse
"""
@Author: LYG
@Date: 2024/11/4
@Description: 
"""
from openpyxl import load_workbook
from logsOutput import set_logger
import logging
logger_debug = set_logger(log_name='ExLogger', name='excel_debug', log_file="excel_debug.log", level=logging.DEBUG)
logger_info = set_logger(log_name='ExLogger', name='excel_info', log_file="excel_info.log", level=logging.INFO)


def execl_qc(province, city_name, map_type):
    # 加载 Excel 文件和工作表
    workbook = load_workbook(f'{province}/{city_name}_{map_type}.xlsx')
    worksheet = workbook.active
    # 使用 set 收集唯一的行数据
    unique_rows = set()
    rows_to_keep = []

    # 逐行遍历工作表中的数据
    for row in worksheet.iter_rows(values_only=True):
        # 将行转换为元组（元组是可哈希的，可以用于 set）
        row_tuple = tuple(row)
        if row_tuple not in unique_rows:
            unique_rows.add(row_tuple)
            rows_to_keep.append(row)  # 仅保存唯一行

    # 清空工作表
    worksheet.delete_rows(1, worksheet.max_row)  # 删除除标题外的所有行

    # 将去重后的行数据写回到工作表
    for row in rows_to_keep:
        worksheet.append(row)

    # 保存去重后的 Excel 文件
    workbook.save(f'{province}/{city_name}_{map_type}.xlsx')
    logger_info.info("去重完成！")
